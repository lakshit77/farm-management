"""Generate an Excel sheet listing the fields used by this platform.

This script produces a human-readable `.xlsx` file that groups the fields we use
into sections (Horse / Rider / Class / Show / Ring / Entry + Results), with basic
formatting suitable for sharing with clients and integration partners.

It intentionally avoids naming any external API vendor in the workbook.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Optional, Sequence, Tuple

from openpyxl import Workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


@dataclass(frozen=True)
class FieldRow:
    """A single row in a section: the field name only."""

    field: str
    description: str


@dataclass(frozen=True)
class Section:
    """A logical group of fields to present together in the workbook."""

    title: str
    rows: Tuple[FieldRow, ...]


def _set_column_widths(ws, widths: Sequence[Tuple[int, float]]) -> None:
    """Apply column widths to a worksheet."""

    for col_idx, width in widths:
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _apply_border(cell: Cell) -> None:
    """Apply a thin grid border to a cell."""

    thin = Side(style="thin", color="D0D0D0")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def _write_title(ws, title: str, subtitle: str) -> int:
    """Write the workbook title block and return the next row index."""

    ws["A1"] = title
    ws["A1"].font = Font(size=16, bold=True, color="111827")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")

    ws["A2"] = subtitle
    ws["A2"].font = Font(size=11, color="374151")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    ws.merge_cells("A1:D1")
    ws.merge_cells("A2:D2")

    return 4


def _write_section(ws, start_row: int, section: Section) -> int:
    """Write one section (header + table) and return the next row index."""

    header_fill = PatternFill("solid", fgColor="EEF2FF")
    header_font = Font(bold=True, color="1E3A8A")

    ws[f"A{start_row}"] = f"For the {section.title}, we use this information"
    ws[f"A{start_row}"].font = Font(bold=True, size=12, color="111827")
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=2)
    ws[f"A{start_row}"].alignment = Alignment(horizontal="left", vertical="center")

    table_header_row = start_row + 1
    ws[f"A{table_header_row}"] = "Field"
    ws[f"B{table_header_row}"] = "Description"

    for col in range(1, 3):
        c = ws.cell(row=table_header_row, column=col)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="left", vertical="center")
        _apply_border(c)

    r = table_header_row + 1
    for row in section.rows:
        ws[f"A{r}"] = row.field
        ws[f"B{r}"] = row.description

        ws[f"A{r}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws[f"B{r}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

        for col in range(1, 3):
            _apply_border(ws.cell(row=r, column=col))

        r += 1

    return r + 1


def _sections() -> Tuple[Section, ...]:
    """Return the ordered sections and their rows (fields used by the platform)."""

    return (
        Section(
            title="horse",
            rows=(
                FieldRow("name", "Horse name."),
                FieldRow("status", "Horse status in our system (for example: active)."),
            ),
        ),
        Section(
            title="rider",
            rows=(
                FieldRow("name", "Rider name."),
            ),
        ),
        Section(
            title="class",
            rows=(
                FieldRow("name", "Class name (the competition class title)."),
                FieldRow("class_number", "Class number shown in the schedule/catalog."),
                FieldRow("sponsor", "Sponsor name for the class (if available)."),
                FieldRow("prize_money", "Prize money amount for the class (if available)."),
                FieldRow("class_type", "Class type/category (for example: jumper, hunter)."),
            ),
        ),
        Section(
            title="show",
            rows=(
                FieldRow("show_id", "Unique identifier for the show."),
                FieldRow("show_name", "Show name."),
                FieldRow("start_date", "Show start date."),
                FieldRow("end_date", "Show end date."),
            ),
        ),
        Section(
            title="ring",
            rows=(
                FieldRow("ring_name", "Ring name (arena name)."),
                FieldRow("ring_number", "Ring number used in the schedule."),
                FieldRow("ring_status", "Ring status for the day (used to determine completion)."),
            ),
        ),
        Section(
            title="entry (per horse per class)",
            rows=(
                FieldRow("entry_id", "Unique identifier for the entry (horse’s participation record)."),
                FieldRow("horse_id", "Identifier for the horse on the source system."),
                FieldRow("rider_id", "Identifier for the rider on the source system (if available)."),
                FieldRow("class_id", "Identifier for the class on the source system."),
                FieldRow("ring_id", "Identifier/number for the ring on the source system."),
                FieldRow("trainer_id", "Identifier for the trainer on the source system (if available)."),
                FieldRow("back_number", "Back number worn by the horse."),
                FieldRow("scheduled_date", "Date the class is scheduled on."),
                FieldRow("estimated_start", "Estimated start time for the class."),
                FieldRow("actual_start", "Actual start time for the class (if available)."),
                FieldRow("order_of_go", "Order in which the horse is expected to go."),
                FieldRow("order_total", "Total number of trips/entries expected in the class (if available)."),
                FieldRow("status", "Entry status in our system (active, completed, scratched)."),
                FieldRow("gone_in", "Whether the horse has completed its round/trip."),
                FieldRow("scratch_trip", "Whether the horse was scratched from the class."),
                FieldRow("class_status", "Overall class status (for example: Not Started, Underway, Completed)."),
                FieldRow("total_trips", "Total trips in the class (if available)."),
                FieldRow("completed_trips", "Trips completed so far."),
                FieldRow("remaining_trips", "Trips remaining."),
            ),
        ),
        Section(
            title="results (per entry)",
            rows=(
                FieldRow("trip_id", "Unique identifier for the trip/result record."),
                FieldRow("placing", "Final placing/position for the entry (if available)."),
                FieldRow("points_earned", "Points earned (if applicable)."),
                FieldRow("total_prize_money", "Prize money earned for the entry (if available)."),
                FieldRow("faults_one", "Faults/penalties for round 1 (if applicable)."),
                FieldRow("time_one", "Time for round 1 (if applicable)."),
                FieldRow("time_fault_one", "Time faults for round 1 (if applicable)."),
                FieldRow("disqualify_status_one", "Disqualification status for round 1 (if applicable)."),
                FieldRow("faults_two", "Faults/penalties for round 2 / jump-off (if applicable)."),
                FieldRow("time_two", "Time for round 2 / jump-off (if applicable)."),
                FieldRow("time_fault_two", "Time faults for round 2 / jump-off (if applicable)."),
                FieldRow("disqualify_status_two", "Disqualification status for round 2 (if applicable)."),
                FieldRow("score1", "Judge score 1 (hunter/scored classes, if applicable)."),
                FieldRow("score2", "Judge score 2 (if applicable)."),
                FieldRow("score3", "Judge score 3 (if applicable)."),
                FieldRow("score4", "Judge score 4 (if applicable)."),
                FieldRow("score5", "Judge score 5 (if applicable)."),
                FieldRow("score6", "Judge score 6 (if applicable)."),
            ),
        ),
    )


def build_workbook() -> Workbook:
    """Create the workbook with formatting and all sections."""

    wb = Workbook()
    ws = wb.active
    ws.title = "Used Fields"

    _set_column_widths(ws, widths=((1, 26.0), (2, 70.0)))

    row = _write_title(
        ws,
        title="Used Fields (Integration Reference)",
        subtitle=(
            "This sheet lists the fields that our platform currently uses.\n"
            "It is grouped by entity for easy comparison with other platforms."
        ),
    )

    for section in _sections():
        row = _write_section(ws, row, section)

    # Freeze header/title area and enable filters only on the first table header row
    ws.freeze_panes = "A5"

    # Light background for the sheet
    ws.sheet_view.showGridLines = False
    ws["A1"].fill = PatternFill("solid", fgColor="FFFFFF")

    return wb


def write_excel(out_path: Path) -> Path:
    """Generate and write the Excel workbook to disk."""

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = build_workbook()
    wb.save(out_path)
    return out_path


def main(argv: Optional[Sequence[str]] = None) -> int:
    """CLI entrypoint."""

    _ = argv  # reserved for future flags
    repo_root = Path(__file__).resolve().parents[1]
    out_path = repo_root / "docs" / "used_fields.xlsx"
    write_excel(out_path)
    print(str(out_path))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

